from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
driver=webdriver.Chrome(ChromeDriverManager().install())

wb = load_workbook('All data of gcp.xlsx')
ws = wb.active
column_b=ws['B'] 

urllist=[]

for cell in column_b:
    website= cell.value
    urllist.append(website)

no=0

for p in range(len(urllist)):
    driver.get(urllist[no])
    driver.implicitly_wait(5)

    badges=driver.find_elements(By.XPATH,"//span[contains(@class,'ql-subhead-1 l-mts')]")


    for badge in badges:

        if badge.text == "Introduction to Digital Transformation with Google Cloud":
            a=1
        if badge.text == "Innovating with Data and Google Cloud":
            b=1
        if badge.text == "Infrastructure and Application Modernization with Google Cloud":
            c=1
        if badge.text == "Understanding Google Cloud Security and Operations":
            d=1

    if a==1 and b==1 and c==1 and d==1:
        column_c=ws.cell(row = no+1, column = 3) 
        column_c.value="All Module Cleared"   
    print(no) 

    if no%10 == 0:
        wb.save(str(no)+"good.xlsx")
        
    no = no+1
    a=0
    b=0
    c=0
    d=0

print("here")
wb.save("finalgood.xlsx")


driver.quit()
